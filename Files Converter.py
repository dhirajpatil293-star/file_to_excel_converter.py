import io
import os
import re
import pandas as pd
import pdfplumber
from PIL import Image
import pytesseract
import streamlit as st

# Configure Tesseract path depending on OS environment
if os.name == "nt":  # Windows local machine
    windows_tesseract = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if os.path.exists(windows_tesseract):
        pytesseract.pytesseract.tesseract_cmd = windows_tesseract

# Page Configuration
st.set_page_config(
    page_title="File to Excel Converter",
    page_icon="⚡",
    layout="wide"
)

# Custom CSS for Developer Branding Footer
st.markdown("""
    <style>
    .developer-footer {
        position: fixed;
        left: 0;
        bottom: 0;
        width: 100%;
        background-color: #0e1117;
        color: #808495;
        text-align: center;
        padding: 8px 0;
        font-size: 14px;
        font-weight: 500;
        border-top: 1px solid #262730;
        z-index: 9999;
    }
    .developer-footer span {
        color: #00d47e;
        font-weight: bold;
    }
    </style>
""", unsafe_allow_html=True)

# Sidebar Developer Info
with st.sidebar:
    st.title("📌 App Info")
    st.write("Convert PDFs, TXT, CSV, and Image files into Excel spreadsheets instantly.")
    st.divider()
    st.caption("💻 Developed by **Dhiraj Patil**")

# Main Header
st.title("⚡ File-to-Excel Converter Web App")
st.write("Upload a **PDF**, **TXT**, **CSV**, or **Image (JPG/PNG)** file to parse its content and export it directly as an Excel spreadsheet.")

# Helper function to extract data
def extract_data_from_file(uploaded_file):
    filename = uploaded_file.name.lower()
    parsed_rows = []
    raw_lines = []

    try:
        if filename.endswith(".pdf"):
            with pdfplumber.open(uploaded_file) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    if tables:
                        for table in tables:
                            for row in table:
                                clean_row = [str(c).replace("\n", " ").strip() for c in row if c is not None]
                                if any(clean_row):
                                    parsed_rows.append(clean_row)
                    else:
                        text = page.extract_text()
                        if text:
                            raw_lines.extend([l.strip() for l in text.split("\n") if l.strip()])

        elif filename.endswith((".txt", ".csv")):
            stringio = io.StringIO(uploaded_file.getvalue().decode("utf-8", errors="ignore"))
            for line in stringio:
                line_str = line.strip()
                if line_str:
                    raw_lines.append(line_str)
                    if "," in line_str:
                        parsed_rows.append([t.strip() for t in line_str.split(",")])
                    elif "\t" in line_str:
                        parsed_rows.append([t.strip() for t in line_str.split("\t")])

        elif filename.endswith((".jpg", ".jpeg", ".png")):
            image = Image.open(uploaded_file)
            text = pytesseract.image_to_string(image)
            raw_lines = [l.strip() for l in text.split("\n") if l.strip()]

    except Exception as e:
        st.error(f"Error processing file: {str(e)}")

    return parsed_rows, raw_lines

def smart_parse_line(line):
    if "," in line:
        return [t.strip() for t in line.split(",") if t.strip()]
    if "\t" in line:
        return [t.strip() for t in line.split("\t") if t.strip()]

    pattern = r"^(\S+)\s+(.+?)\s+([A-Za-z0-9\s]+?)\s+(\d+)\s+([\d\.]+)\s+(.+)$"
    match = re.match(pattern, line)
    if match:
        return [g.strip() for g in match.groups()]

    tokens = [t.strip() for t in re.split(r"\s{2,}", line) if t.strip()]
    if len(tokens) > 1:
        return tokens

    return line.split()

# File Uploader
uploaded_file = st.file_uploader("Choose a file to convert", type=["pdf", "txt", "csv", "jpg", "jpeg", "png"])

if uploaded_file is not None:
    parsed_rows, raw_lines = extract_data_from_file(uploaded_file)

    if not parsed_rows and not raw_lines:
        st.error("Could not extract readable text or table data from this file.")
    else:
        st.subheader("1. Extracted Data Preview")
        default_headers_str = "Part ID, Component Name, Material Grade, Quantity, Unit Price ($), Status"

        if parsed_rows:
            st.dataframe(pd.DataFrame(parsed_rows[:10]), use_container_width=True)
            if any("part" in str(c).lower() for c in parsed_rows[0]):
                default_headers_str = ", ".join(parsed_rows[0])
        else:
            st.code("\n".join(raw_lines[:10]), language="text")
            detected = smart_parse_line(raw_lines[0])
            if len(detected) > 1:
                default_headers_str = ", ".join(detected)

        st.subheader("2. Configure Excel Headers")
        headers_input = st.text_input("Enter column header names separated by commas:", value=default_headers_str)

        custom_headers = [h.strip() for h in headers_input.split(",") if h.strip()]

        if custom_headers:
            formatted_rows = []

            if parsed_rows:
                source_rows = parsed_rows
                if any("part" in str(cell).lower() for cell in source_rows[0]):
                    source_rows = source_rows[1:]

                for row in source_rows:
                    row_dict = {}
                    for idx, header in enumerate(custom_headers):
                        row_dict[header] = row[idx] if idx < len(row) else ""
                    formatted_rows.append(row_dict)

            elif raw_lines:
                source_lines = raw_lines
                if any(k in source_lines[0].lower() for k in ["part", "item", "component"]):
                    source_lines = source_lines[1:]

                for line in source_lines:
                    tokens = smart_parse_line(line)
                    row_dict = {}
                    for idx, header in enumerate(custom_headers):
                        row_dict[header] = tokens[idx] if idx < len(tokens) else ""
                    formatted_rows.append(row_dict)

            df = pd.DataFrame(formatted_rows, columns=custom_headers)

            st.subheader("3. Structured Excel Preview")
            st.dataframe(df, use_container_width=True)

            output_buffer = io.BytesIO()
            with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Converted Data", index=False)
                ws = writer.sheets["Converted Data"]

                from openpyxl.styles import Font, PatternFill
                header_fill = PatternFill(start_color="00472B", end_color="00472B", fill_type="solid")
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                for col in ws.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

            st.download_button(
                label="📥 Download Excel File",
                data=output_buffer.getvalue(),
                file_name="converted_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# Sticky Bottom Footer for Developer Branding
st.markdown(
    '<div class="developer-footer">Developed by <span>Dhiraj Patil</span></div>',
    unsafe_allow_html=True
)
